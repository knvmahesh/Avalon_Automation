import streamlit as st
from mcp.server.fastmcp import FastMCP
import asyncio
import json
import nest_asyncio
import time
import os
from datetime import datetime
from typing import Any, Dict, Optional
from dotenv import find_dotenv, load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException,NoSuchElementException
import re
import Excel
import openpyxl
import messagebox


# MCP clients
from mcp import ClientSession, StdioServerParameters
from mcp.client.stdio import stdio_client
from mcp.client.sse import sse_client

mcp = FastMCP(name="Fusion_Automation")
print("FastMCP initialized.")

###Loading configurations from .env file   
dotenv_path = find_dotenv()
print(".env path:", dotenv_path)
if dotenv_path:
    load_dotenv(dotenv_path)
else:
    # last-resort: attempt .env next to this module
    load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

excel_path = os.getenv("EXCEL_PATH")
print("Excel Path from .env:", excel_path)
if not excel_path:
    #st.warning("EXCEL_PATH not found in .env or environment variables. Set EXCEL_PATH=full\\path\\to\\file.xlsx in your .env (no surrounding quotes) or export EXCEL_PATH in the environment.")
    # optional fallback so the rest of the code doesn't immediately break
    excel_path = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"

    print(f"Excel Path resolved: {excel_path}")   


nest_asyncio.apply()

st.set_page_config(page_title="MCP Chat", layout="wide")

#st.title("Avalon Automation Tool")
st.markdown("<h1 style='text-align: center;'>Avalon Automation Tool</h1>", unsafe_allow_html=True)
# Transport settings (set a default)
transport = "sse"

# --- Inject Custom CSS for Cancel Button Color (UNCHANGED) ---
# Target the 'Cancel' button using its text content (label) in CSS.
# Note: The 'Submit' button color is still handled by .streamlit/config.toml
CANCEL_BUTTON_COLOR = "#0080ff"
CSS_HACK = f"""
<style>
/* Target the button based on its content (Cancel) */
div.stButton > button:nth-child(1) {{
    background-color: {CANCEL_BUTTON_COLOR};
    color: white; /* Ensure text is readable */
    border-color: {CANCEL_BUTTON_COLOR};
}}
/* Apply hover/focus state for better UX */
div.stButton > button:nth-child(1):hover, 
div.stButton > button:nth-child(1):focus {{
    background-color: #0055aa; /* Slightly darker shade for hover */
    border-color: #0055aa; 
}}
</style>
"""
st.markdown(CSS_HACK, unsafe_allow_html=True)
# --- End Custom CSS Injection ---


def small_info(text: str, font_size: str = '14px', bgcolor: str = '#f0f8ff', border_color: str = '#00aaff', padding: str = '12px', width: str = '600px'):
    """Display a compact info-style message with reduced font size, padding, and optional width.

    Parameters:
        text: The message text to display.
        font_size: CSS font-size (e.g., '12px' or '0.9rem').
        bgcolor: Background color of the info box.
        border_color: Color of the left border accent.
        padding: Inner padding for the box.
        width: Optional CSS width (e.g., '300px', '50%', 'auto').
    """
    # Use max-width:100% to avoid overflow when a fixed width is provided
    html = f"""<div style="background-color:{bgcolor}; border-left: 4px solid {border_color}; padding:{padding}; border-radius:4px; font-size:{font_size}; margin: 8px 0; width:{width}; max-width:100%;">{text}</div>"""
    st.markdown(html, unsafe_allow_html=True)

# --- UI Configuration (Dropdowns) ---

# Use columns to restrict the width of the dropdowns
# col_dropdown will be 4 parts wide, and col_space will be 6 parts wide (40% width)
col_dropdown, col_space = st.columns([4, 6])

with col_dropdown:
    # 1. Initial Selectbox (ToolType) - Now restricted by col_dropdown width
    toolType = st.selectbox("ToolType", ("Select Tool", "RWB", "Run Automation Scripts"))

# Define variables with placeholders to ensure they exist for button handlers
tool_name = ""
args_json = "{}"
sse_url = ""
proxy_token = ""

# 2. Logic to handle the selection
if toolType == "RWB":
    #st.header("RWB Configuration")
    #sse_url = st.text_input("SSE URL", value="http://localhost:8050/sse")
    #proxy_token = st.text_input("Proxy token (optional)", value="", type="password")
    
    #tool_name = st.text_input("Tool name (exact)", value="add")
    args_json = st.text_area("Tool arguments (JSON)", value='{"a":2,"b":3}', height=120)

elif toolType == "Run Automation Scripts":
    # 1. Module Selection
    with col_dropdown:
        selected_module = st.selectbox("Select Module", ["Select Module", "HCM", "SCM"])

    # Define the mapping
    script_mapping = {
        "HCM": [
            "Hire to Fire",
            "Department creation",
            "Job creation",
            "Position creation",
            "Employee creation",
            "Employee Termination",
            "Business Unit Creation"
        ],
        "SCM": [
            "Purchase Order Creation",
            "Inventory Update"
        ]
    }

    if selected_module == "Select Module":
        with col_dropdown:
            st.selectbox("Select Script", ["Select Module first"], disabled=True)
        small_info("Please choose a Module to view available automation scripts.")
        tool_name = ""
    
    else:
        # 2. Script Selection (Visible once Module is selected)
        automation_options = ["Select Automation Script"] + script_mapping.get(selected_module, [])
        
        with col_dropdown:
            selected_script = st.selectbox("Select Script", automation_options)
        
        # 3. Message logic: Displayed after Module is picked but before Script is picked
        if selected_script == "Select Automation Script":
            small_info(f"Please select a script from the **{selected_module}** module to proceed.")
            tool_name = ""
            args_json = "{}"
        else:
            # Script has been selected
            tool_name = selected_script.replace(" ", "_")
            small_info(
                text=f"Ready to run: **{selected_script}**", 
                bgcolor="#e6ffed", 
                border_color="#28a745"
            )
            #st.success(f"Ready to run: {selected_script}")
elif toolType == "Select Tool":
    small_info("Please choose a tool (RWB or Run Automation Scripts) to view configuration options.")


# --- Chat State and Button Placement (Colored and Sized) ---

# Chat state
if "chat" not in st.session_state:
    st.session_state.chat = []

# Use a ratio for st.columns to shrink the buttons (ratio: 1, 1, 1, 4)
col_space_l, col_submit, col_cancel, col_space_r = st.columns([1, 1, 1, 4]) 

with col_submit:
    # Submit remains 'primary'. It gets its custom color #00c8ff from config.toml.
    call_button = st.button(
        "Submit", 
        type="primary", 
        use_container_width=True
    )

with col_cancel:
    # Cancel is set to 'secondary'. The CSS hack above overrides this button's color.
    list_button = st.button(
        "Cancel", 
        type="secondary", 
        use_container_width=True
    )

# --- Helpers: async call wrappers (UNCHANGED) ---
async def _call_tool_sse(url: str, token: Optional[str], tool: str, args: Dict[str, Any]):
    headers = {"Authorization": f"Bearer {token}"} if token else None
    async with sse_client(url, headers=headers) as (read_stream, write_stream):
        async with ClientSession(read_stream, write_stream) as session:
            await session.initialize()
            return await session.call_tool(tool, arguments=args)

async def _call_tool_stdio(server_script_path: str, tool: str, args: Dict[str, Any]):
    params = StdioServerParameters(command="python", args=[server_script_path])
    async with stdio_client(params) as (read_stream, write_stream):
        async with ClientSession(read_stream, write_stream) as session:
            await session.initialize()
            return await session.call_tool(tool, arguments=args)

async def _list_tools_sse(url: str, token: Optional[str]):
    headers = {"Authorization": f"Bearer {token}"} if token else None
    async with sse_client(url, headers=headers) as (r, w):
        async with ClientSession(r, w) as session:
            await session.initialize()
            return await session.list_tools()

async def _list_tools_stdio(server_script_path: str):
    params = StdioServerParameters(command="python", args=[server_script_path])
    async with stdio_client(params) as (r, w):
        async with ClientSession(r, w) as session:
            await session.initialize()
            return await session.list_tools()

# --- Helper Functions for Excel Operations ---
def getRowCount(path: str, sheet_name: str) -> int:
    """Get the number of rows in an Excel sheet"""
    import openpyxl
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row

def readData(path: str, sheet_name: str, row: int, col: int) -> str:
    """Read data from a specific cell in an Excel sheet"""
    import openpyxl
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row, column=col).value

def writeData(path: str, sheet_name: str, row: int, col: int, value: str) -> None:
    """Write data to a specific cell in an Excel sheet"""
    import openpyxl
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row, column=col).value = value
    workbook.save(path)

def do_list_tools():
    try:
        if transport == "sse" and sse_url:
            # result = asyncio.run(_list_tools_sse(sse_url, proxy_token or None)) # Uncomment for actual functionality
            result = type('ToolList', (object,), {'tools': [type('Tool', (object,), {'name': 'tool1', 'description': 'desc1'}), type('Tool', (object,), {'name': 'tool2', 'description': 'desc2'})]})
        else:
            result = None
            st.session_state.chat.append({"role": "server", "text": "Connection transport not fully configured."})
            return

        tools = getattr(result, "tools", [])
        lines = [f"{t.name} - {getattr(t, 'description', '')}" for t in tools]
        st.session_state.chat.append({"role": "server", "text": "Available tools:\n" + "\n".join(lines)})
    except Exception as e:
        st.session_state.chat.append({"role": "server", "text": f"Error listing tools: {e}"})

@mcp.tool()        
def Manage_Departments (path: str = excel_path):

    """Department Creation"""
    messagebox.showinfo("Info", "Department Creation Automation Script Started Successfully.")  
    excel_exists = Excel.check_excel_exists(path)
    if not excel_exists:
        st.error("Excel file not found. Please check the file path.")
        return
    Excel.close_open_excel()    
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
    service=ChromeService(ChromeDriverManager().install()), 
    options=options
        
        )
    driver.maximize_window()
          
    #path = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    #rows1 = getRowCount(path,'Config')
    rows2 = getRowCount(path,'Dept Creation')
    Base_url = readData(path,"Config",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"Config",2,2)
    PW = readData(path,"Config",2,3)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()
    
    # - clicking on home button
    
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()
    
    # - clicking on my client groups
    
    xpath_MCG = '//a[@id="groupNode_workforce_management"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()
    
    # - clicking on workforce structure
    
    xpath_WS = '//*[@id="itemNode_workforce_management_workforce_structures_0"]'
    WSClick = WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH, xpath_WS)))
    WSClick.click()
    
    # - searching for manage department and clicking on it
    
 
    
    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Manage Departments")))
    link = driver.find_element(By.PARTIAL_LINK_TEXT, "Manage Departments")
    link.click()
    
    for r in range (2,rows2+1):
        xpath_create = "//a[./span[text()='Create']]"
        createClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_create)))
        createClick.click()
    
        xpath_calender = "//input[@aria-label='Effective Start Date']"
        searchBox = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_calender)))
        searchBox.clear()
        date = readData(path,"Dept Creation",r,1)
        date_string = date.strftime('%d-%b-%Y')
        searchBox.send_keys(date_string)
    
        xpath_deptSelect = "//a[@title='Search: Department Set']"
        deptSelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_deptSelect)))
        deptSelectClick.click()
    
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
    
        xpath_RDSN = "//input[@aria-label=' Reference Data Set Name']"
        RDSN_SB = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_RDSN)))
        RDSN_SB.clear()
        RDS = readData(path,"Dept Creation",r,2)
        RDSN_SB.send_keys(RDS+ Keys.ENTER)
    
        xpath_selectRDSN = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:AP1:setName2Id_afrLovInternalTableId::db"]/table/tbody/tr/td[1]'
        selectRDSN = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectRDSN)))
        selectRDSN.click()
    
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
        time.sleep(2)
       
        NM = readData(path,"Dept Creation",r,3)       
        driver.switch_to.active_element.send_keys(Keys.TAB + NM)
        xpath_statusDD = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:AP1:selectOneChoice2::content"]'
        statusDD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_statusDD)))
        statusDD.click()
    
        xpath_Astatus = "//li[text()='Active']"
   
        # 1. Click 'Active' status (This action triggers the Name Exists validation)
        Astatus = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_Astatus)))
        Astatus.click()
        time.sleep(2)
 
        # - START DIALOG CHECK BLOCK
        # - We use a try/except to check for the MANDATORY presence of the error dialog button (OK)
        # - If the OK button is found, we assume the error occurred and execute the 'except' block.
   
        xpath_error_ok_button = "//button[text()='OK']"
   
        try:    
       
            error_ok_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_error_ok_button)))
            error_ok_button.click()
            time.sleep(1)
 
            xpath_clickCE = "//a[normalize-space(.)='Cancel']"
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_clickCE))).click()
            time.sleep(1)
 
            xpath_clickYE_confirmation = "//button[@accesskey='Y']"
            try:
                #- If it appears, clicks it.
                WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, xpath_clickYE_confirmation))).click()
                time.sleep(1)
            except TimeoutException:
                pass
            writeData(path,"Dept Creation",r,4,"Fail")
            writeData(path,"Dept Creation",r,5,"NAME Already Exists")
            time.sleep(3)
            
            continue
           
        except TimeoutException:
        #- IF DIALOG DID NOT APPEAR
               
            xpath_clickNext = "//a[normalize-space(.)='Next']"
            clickNext = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickNext)))
            clickNext.click()
 
            xpath_clickNext2 = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:3:AP2:tt1:next"]/a/span'
            clickNext2 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickNext2)))
            clickNext2.click()
 
            xpath_clickSubmit = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:4:ap1:tt1:submit"]/a/span'
            clickSubmit = WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickSubmit)))
            clickSubmit.click()
 
            xpath_clickYes = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:4:ap1:tt1:okWarningDialog"]'
            clickYes = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickYes)))
            clickYes.click()
 
            xpath_clickOK = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:4:ap1:tt1:okConfirmationDialog"]'
            clickOK = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickOK)))
            clickOK.click()
            writeData(path,"Dept Creation",r,4,"Pass")
            writeData(path,"Dept Creation",r,5,"Department Created Successfully")
            time.sleep(2)
        
    driver.quit()
    wb = openpyxl.load_workbook(path)
    excel_bytes = Excel.workbook_to_bytes(wb)
    
    st.download_button(
                 label="Download Excel File",
                 data = excel_bytes,
                file_name=f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    messagebox.showinfo("Success", path)   
    #md_Summary=Excel.summarysheet_markdown(path)
    #messagebox.showinfo("Success", md_Summary)
    #print(md_Summary)
    return path
@mcp.tool()
def Manage_Positions (path: str = excel_path):
    
    excel_exists = Excel.check_excel_exists(path)
    if not excel_exists:
        st.error("Excel file not found. Please check the file path.")
        return
    Excel.close_open_excel()
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), 
    options=options  )
    driver.maximize_window()
    rows_Termination = getRowCount(path, 'Pos Creation')
          
    #path = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    #rows1 = getRowCount(path,'Config')
    rows2 = getRowCount(path,'Pos Creation')
    Base_url = readData(path,"Config",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"Config",2,2)
    PW = readData(path,"Config",2,3)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()   

    # - clicking on home button

    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()

    # - clicking on settings&actions, then clicking on setup & maintenance, clicking on taskbar

    xpath_ic = "//img[@title='Settings and Actions']"
    ic = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ic)))
    ic.click()

    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Setup and Maintenance")))
    SMlink = driver.find_element(By.PARTIAL_LINK_TEXT, "Setup and Maintenance")
    SMlink.click()

    xpath_Tasks = "//img[@title='Tasks']"
    Tasks = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Tasks)))
    Tasks.click()

    # - Clicking on Search -> searching managa positions and clicking on it

    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
    Searchlink1 = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
    Searchlink1.click()

    xpath_ITF = "//input[contains(@id, ':s9:it1::content')]"
    ITF = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ITF)))
    ITF.click()
    ITF.send_keys("MANAGE POSITIONS"+ Keys.ENTER)

    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Manage Positions")))
    MPLink = driver.find_element(By.PARTIAL_LINK_TEXT, "Manage Positions")
    MPLink.click()

    #xpath_ADD = "//input[contains(@id, ':s9:it1::content')]"
    #ITF = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ITF)))
    #ITF.click()

    # - clicking on ADD and entering details in 1st BASIC DETAILS page
    for r in range (2,rows2+1):

        xpath_ADD = "//img[contains(@src, 'func_add_16_ena.png')]"
        ADD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ADD)))
        ADD.click()

        xpath_ESD = "//input[@aria-label='Effective Start Date']"
        ESD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ESD)))
        ESD.click()
        ESD.clear()
        ESDe = readData(path,"Pos Creation",r,1)
        #date_string = ESDe.strftime('%d-%b-%Y')
        date_string = ESDe.strftime('%m/%d/%y')
        ESD.send_keys(date_string)

        xpath_BU = "//input[contains(@id, ':AP4:BUNameId::content')]"
        BU = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_BU)))
        BU.click()
        BUe = readData(path,"Pos Creation",r,2)
        BU.send_keys(BUe+Keys.ENTER)
        time.sleep(2)

        xpath_name1 = "//input[contains(@id, ':AP4:inputText2::content')]"
        name1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_name1)))
        name1.click()
        nameE = readData(path,"Pos Creation",r,3)
        name1.send_keys(nameE+Keys.ENTER)
        time.sleep(2)

        xpath_next1 = "//a[@accesskey='x']"
        next1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next1)))
        next1.click()

        # - entering details POSITION DETAILS page

        xpath_AStatus = "//input[@title='Active']"
        AStatus = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_AStatus)))
        AStatus.click()
        AStatusE = readData(path,"Pos Creation",r,4)
        AStatus.send_keys(AStatusE+Keys.ENTER)

        xpath_deptDD = "//a[@title='Search: Department']"
        deptDD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_deptDD)))
        deptDD.click()
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        SLink1 = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        SLink1.click()

        xpath_deptN = "//input[@aria-label=' Name']"
        deptN = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_deptN)))
        deptN.clear()
        deptNE = readData(path,"Pos Creation",r,5)
        deptN.send_keys(deptNE+Keys.ENTER)
        time.sleep(3)
        xpath_selectDept = "//td[@style='width:6px;' and contains(@class, 'xwn')]"
        selectDept = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectDept)))
        selectDept.click()
        xpath_deptOK = "//button[text()='OK']"
        deptOK = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_deptOK)))
        deptOK.click()
        time.sleep(3)

        xpath_jobDD = "//a[@title='Search: Job']"
        jobDD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobDD)))
        jobDD.click()
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        SLink1 = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        SLink1.click()
        xpath_jobName = "//input[@aria-label=' Name']"
        jobName = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobName)))
        jobName.clear()
        jobNameE = readData(path,"Pos Creation",r,6)
        jobName.send_keys(jobNameE+Keys.ENTER)
        time.sleep(2)
        xpath_jobSelect = "//td[@style='width:6px;' and contains(@class, 'xwn')]"
        jobSelect = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobSelect)))
        jobSelect.click()
        xpath_jobOK = "//button[text()='OK']"
        jobOK = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobOK)))
        jobOK.click()
        time.sleep(2)


        xpath_HS = "//input[contains(@id, ':selectOneChoice4::content') and @class='x2h']"
        HS = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_HS)))
        HS.click()
        HSe = readData(path,"Pos Creation",r,7)
        HS.send_keys(HSe+Keys.ENTER)
        time.sleep(2)

        xpath_type = "//a[contains(@id, ':selectOneChoice5::drop') and @class='x1kt']"
        type = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_type)))
        type.click()
        typeE = readData(path,"Pos Creation",r,8)
        type.send_keys(typeE+Keys.ENTER)
        time.sleep(2)

        xpath_HC = "//input[contains(@id, ':AP5:inputText7::content')]"
        HC = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_HC)))
        HC.clear()
        HCe = readData(path,"Pos Creation",r,9)
        HC.send_keys(HCe)
        time.sleep(2)

        xpath_next2 = "//a[normalize-space(.)='Next']"
        next2 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next2)))
        next2.click()
        time.sleep(2)
        xpath_next3 = "//a[normalize-space(.)='Next']"
        next3 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next3)))
        next3.click()
        time.sleep(2)

        xpath_next4 = "//a[normalize-space(.)='Next']"
        next4 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next4)))
        next4.click()
        time.sleep(2)

        xpath_submit = "//a[normalize-space(.)='Submit']"
        submit = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_submit)))
        submit.click()
        time.sleep(2)

        xpath_yes1 = "//button[@accesskey='Y']"
        yes1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_yes1)))
        yes1.click()
        time.sleep(5)

        driver.switch_to.active_element.send_keys(Keys.TAB)
        driver.switch_to.active_element.send_keys(Keys.ENTER)

        writeData(path,"Pos Creation",r,10,"Pass")
        writeData(path,"Pos Creation",r,11,"Position Created Successfully")

    time.sleep(3)
    driver.quit() 

    wb = openpyxl.load_workbook(path)
    excel_bytes = Excel.workbook_to_bytes(wb)
    
    st.download_button(
                 label="Download Excel File",
                 data = excel_bytes,
                file_name=f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
@mcp.tool()
def Manage_Jobs (path: str = excel_path):
    """Job Creation"""
    
    excel_exists = Excel.check_excel_exists(path)
    if not excel_exists:
        st.error("Excel file not found. Please check the file path.")
        return  
    Excel.close_open_excel()
    print("Hi1")
    #Browser opening
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), 
    options=options  
    )
    driver.maximize_window()
    #path = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    #rows1 = getRowCount(path,'Config')
    rows2 = getRowCount(path,'Job Creation')
    Base_url = readData(path,"Config",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"Config",2,2)
    PW = readData(path,"Config",2,3)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()
    
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()
    
    # - clicking on settings&actions, then clicking on setup & maintenance, clicking on taskbar
    
    xpath_ic = "//img[@title='Settings and Actions']"
    ic = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ic)))
    ic.click()
    
    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Setup and Maintenance")))
    SMlink = driver.find_element(By.PARTIAL_LINK_TEXT, "Setup and Maintenance")
    SMlink.click()
    
    xpath_Tasks = "//img[@title='Tasks']"
    Tasks = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Tasks)))
    Tasks.click()
    
    # - Clicking on Search -> searching managa job and clicking on it
    
    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
    Searchlink1 = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
    Searchlink1.click()
    
    xpath_ITF = "//input[contains(@id, ':s9:it1::content')]"
    ITF = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ITF)))
    ITF.click()
    ITF.send_keys("MANAGE JOB"+ Keys.ENTER)
    
    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Manage Job")))
    MPLink = driver.find_element(By.PARTIAL_LINK_TEXT, "Manage Job")
    MPLink.click()
    rowcount=0
    for r in range (2,rows2+1):
        xpath_ADD = "//img[contains(@src, 'func_add_16_ena.png')]"
        ADD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ADD)))
        ADD.click()
        rowcount=rowcount+1
        xpath_ESD = "//input[@aria-label='Effective Start Date']"
        ESD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ESD)))
        ESD.click()
        ESD.clear()
        ESDe = readData(path,"Job Creation",r,1)
        #date_string = ESDe.strftime('%d-%b-%Y')
        date_string = ESDe.strftime('%m/%d/%y')
        #12/17/25
        ESD.send_keys(date_string)
        
        xpath_jobDD = "//a[@title='Search: Job Set']"
        jobDD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobDD)))
        jobDD.click()
        element_to_click = WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        driver.execute_script("arguments[0].scrollIntoView();", element_to_click)
        SLink1 = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        SLink1.click()
 
    
        xpath_RDSN = "//input[@aria-label=' Reference Data Set Name']"
        RDSN_SB = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_RDSN)))
        RDSN_SB.clear()
        RDS = readData(path,"Job Creation",r,2)
        RDSN_SB.send_keys(RDS+ Keys.ENTER)
    
        xpath_selectRDSN = '//*[@id="pt1:r1:0:rt:1:r2:0:dynamicRegion1:1:basicDetailsPnl:setName_afrLovInternalTableId::db"]/table/tbody/tr/td[1]'
        selectRDSN = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectRDSN)))
        selectRDSN.click()
    
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
    
        xpath_name = "//input[contains(@id, ':basicDetailsPnl:name::content')]"
        name = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_name)))
        name.clear()
        nameE = readData(path,"Job Creation",r,3)
        name.send_keys(nameE+Keys.ENTER)
    
        xpath_jobcode = "//input[contains(@id, ':basicDetailsPnl:jobCode::content')]"
        jobcode = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobcode)))
        jobcode.clear()
        JCE = readData(path,"Job Creation",r,4)
        jobcode.send_keys(JCE+Keys.ENTER)
        time.sleep(3)
    
        xpath_error_icon = "//td[contains(text(), '(PER-1530038)')]"
        xpath_clickCE = "//a[normalize-space(.)='Cancel']"
        try :
            error_icon = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath_error_icon)))
            cancel_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath_clickCE)))
            driver.execute_script("arguments[0].click();", cancel_button)
            xpath_clickYE_confirmation = "//button[@accesskey='Y']" 
            try:
                #- If it appears, clicks it.
                WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, xpath_clickYE_confirmation))).click()
                time.sleep(1) 
            except TimeoutException:
                pass 
            writeData(path,"Job Creation",r,6,"Fail")
            writeData(path,"Job Creation",r,7,"JobCode already Exists")
            #st.success("Execution completed successfully")
            #st.link_button("Download Execution Summary", result_excel_url)
            time.sleep(3)
            continue
        except TimeoutException:
            pass        
    
        xpath_next1 = "//a[@accesskey='x']"
        next1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next1)))
        next1.click()
        time.sleep(3)
    
        xpath_AStatus = "//input[@title='Active']"
        AStatus = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_AStatus)))
        AStatus.click()
        AStatusE = readData(path,"Job Creation",r,5)
        AStatus.send_keys(AStatusE+ Keys.ENTER)
    
        xpath_next2 = "//a[@accesskey='x']"
        next2 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next2)))
        next2.click()
        time.sleep(3)
    
        xpath_next3 = "//a[@accesskey='x']"
        next3 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next3)))
        next3.click()
        time.sleep(3)
    
        xpath_next4 = "//a[@accesskey='x']"
        next4 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next4)))
        next4.click()
        time.sleep(3)
    
        xpath_submit = "//a[normalize-space(.)='Submit']"
        submit = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_submit)))
        submit.click()
        time.sleep(2)
    
        xpath_yes1 = "//button[@accesskey='Y']"
        yes1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_yes1)))
        yes1.click()
        time.sleep(3)
    
        driver.switch_to.active_element.send_keys(Keys.TAB)
        driver.switch_to.active_element.send_keys(Keys.ENTER)
    
        writeData(path,"Job Creation",r,6,"Pass")
        writeData(path,"Job Creation",r,7,"JOB Created Successfully")
    time.sleep(3)
    

    driver.quit()
    wb = openpyxl.load_workbook(path)
    excel_bytes = Excel.workbook_to_bytes(wb)
    
    st.download_button(
        label="Download Excel File",
        data = excel_bytes,
        file_name=f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
@mcp.tool()
def Employee_Creation (path: str = excel_path):
    """Creation of an Employee"""
    excel_exists = Excel.check_excel_exists(path)
    if not excel_exists:
        st.error("Excel file not found. Please check the file path.")
        return  
    Excel.close_open_excel()
       
# ---------------------------------------------------------------------
#  LAUNCH BROWSER
# ---------------------------------------------------------------------
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()), 
        options=options
        )
    driver.maximize_window()
    
    # ---------------------------------------------------------------------
    #  EXCEL PATH & SHEET ROWS
    # ---------------------------------------------------------------------
   # path = "C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    rows2 = getRowCount(path, 'Emp Creation')
    rows_URL = getRowCount(path, 'Config')
    # ---------------------------------------------------------------------
    #  LOGIN
    # ---------------------------------------------------------------------
    Base_url = readData(path, "Config", 2, 1)
    driver.get(Base_url)
    
    UN = readData(path, "Config", 2, 2)
    PW = readData(path, "Config", 2, 3)
    
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()
    
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()
    
    # - clicking on my client groups
    
    xpath_MCG = '//a[@id="groupNode_workforce_management"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()
    time.sleep(5)
    
    # - Click on New Person 
    xpath_MCG = '//a[@id="itemNode_workforce_management_new_person_0"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()    
    time.sleep(5)
    
    for r in range (2,rows2+1):
        # - Clcik on Hire an Employee
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Hire an Employee")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Hire an Employee")
        link.click()
        
        # Hire Date 
        Hire_Date = readData(path,"Emp Creation",r,1)
        xpath_HireDate = "//input[@aria-label='Hire Date']"
        HireDate = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_HireDate)))
        HireDate.clear()
        #Emp_Hire_Date = Hire_Date.strftime('%d-%b-%Y')
        Emp_Hire_Date = Hire_Date.strftime('%m/%d/%y')
        HireDate.send_keys(Emp_Hire_Date)
        time.sleep(5)
        # Hire Action dropdown
    
        try:
    
            excel_HireAction = readData(path, 'Emp Creation', r, 2)
            # 1. Click the dropdown disclosure icon to open the list
            xpath_HireAction_disclose = '//a[contains(@id, "selectOneChoice1::drop")]'
    
            HireAction_disclose = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_HireAction_disclose))        )
    
            HireAction_disclose.click()
            print("Clicked Hire Action disclosure icon.")
            xpath_List_Item = f"//li[text()='{excel_HireAction}']" 
            HireAction_Option = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_List_Item)))
            HireAction_Option.click()
            print(f"Selected Hire Action: {excel_HireAction}") 
            time.sleep(2) # Allow time for the selection to register and the field to close
    
        except TimeoutException:
            print(f"Error: Hire Action '{excel_HireAction}' not found or clickable for row {r}.")
            #writeData(path, "Employee", r, 5, "Fail - Hire Action Select Timeout")
            continue
    
        except Exception as e:
            # Catch all other exceptions
            print(f"Error setting Hire Action for row {r}: {e}")
            #writeData(path, "Employee", r, 5, "Fail - Hire Action Error")
            continue
    
        # - Click On legal employer tab
        xpath_LESelect = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:selectOneChoice3::lovIconId"]'
        LESelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_LESelect)))
        LESelectClick.click()
        # - Select the search Button
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
        # - Give the legal employer Details
        xpath_LE = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:selectOneChoice3::_afrLovInternalQueryId:value00::content"]'
        Legal_Employer= WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_LE)))
        Legal_Employer.clear()
        LegalEntity= readData(path,"Emp Creation",r,3)
        Legal_Employer.send_keys(LegalEntity+ Keys.ENTER)
        xpath_selectLE = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:selectOneChoice3_afrLovInternalTableId::db"]/table/tbody/tr/td[2]/div/table/tbody/tr/td[1]'
        selectLE = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectLE)))
        selectLE.click()
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
    
        time.sleep(5)
        print("Legal Employer Selected")
        # - Enter Last Name
        Last_Name= readData(path,"Emp Creation",r,4)
        print("Last_Name:", Last_Name)
        #xpath_LastName = "//input[contains(@name, 'i1:4:it20')]"
        xpath_LastName = "//input[contains(@id, 'it20::content') and @type='text']"
        LastName = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,xpath_LastName)))
        LastName.clear()
        #LastName.send_keys(Last_Name+ Keys.TAB)
        LastName.send_keys(Last_Name)
        print("LastName")
    
        # - Enter First Name
        First_Name= readData(path,"Emp Creation",r,5)
        xpath_FristName = '//input[contains(@id, "it60::content") and @type="text"]'
        #xpath_FristName = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:NewPe1:0:pt_r1:0:r1:0:i1:1:it60::content"]'
        FirstName = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,xpath_FristName)))
        FirstName.clear()
        #FirstName.send_keys(First_Name+ Keys.TAB)
        FirstName.send_keys(First_Name)
        print("FirstName")
        
        try:
            # Read Gender from Excel
            excel_Gender = readData(path, 'Emp Creation', r, 6).strip()
                      
            # 1. Click Gender dropdown (disclosure icon)
            xpath_Gender = '//a[contains(@id, "soc3::drop")]'
            
            Gender_dropdown = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, xpath_Gender))
            )
            Gender_dropdown.click()
            print("✅ Clicked Gender dropdown")
            
            # 2. Select Gender value from list
            xpath_GenderSelect = (
                f"//ul[contains(@id,'soc3::pop')]//li[normalize-space()='{excel_Gender}']"
            )
            #<li role="option" class="x1l5" _adfiv="3">Female</li>
            #//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:NewPe1:0:pt_r1:0:soc3::pop"]/li[6]
            Gender_option = WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.XPATH, xpath_GenderSelect))
            )
            Gender_option.click()
            #Gender_option.send_keys(Gender_option)
            print(Gender_option)
            
            print(f"✅ Selected Gender: {excel_Gender}")
            time.sleep(2)

        except TimeoutException:
            print(f"❌ Gender '{excel_Gender}' not found or not clickable for row {r}")
            # writeData(path, "Employee", r, 5, "Fail - Gender Select Timeout")
            continue

        except Exception as e:
            print(f"❌ Error setting Gender for row {r}: {e}")
            # writeData(path, "Employee", r, 5, "Fail - Gender Select Error")
            continue
        
        
        
        # - Choose Gender
        #xpath_Gender= '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:NewPe1:0:pt_r1:0:soc3::drop"]'
        #Gender = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Gender)))
        #Gender.click()
        #time.sleep(5)
        #
        ## - select respective Gender
        #xpath_Gender1 = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:NewPe1:0:pt_r1:0:soc3::pop"]/li[4]'
        #Gender1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Gender1)))
        #Gender1.click()
        #time.sleep(3)
    
        # - Click on Next Page
        xpath_NextPage = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:tt1:next"]/a/span'
        NextPage = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NextPage)))
        NextPage.click()
        time.sleep(3)

        person_label_xpath = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:1:SP1:pt_r1:0:panelLabelAndMessage4"]/td[1]/label'
        person_number_xpath = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:1:SP1:pt_r1:0:panelLabelAndMessage4"]/td[2]'
        
        try:
            # Check if label exists
            label_element = driver.find_element(By.XPATH, person_label_xpath)
            
            # If found, get the person number
            person_number_element = driver.find_element(By.XPATH, person_number_xpath)
            person_number = person_number_element.text.strip()
        
            print("Person Number:", person_number)
        
        except NoSuchElementException:
            print("Person Number label not found")

            # - Click on Next Page
        xpath_NextPage = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:1:SP1:tt1:next"]/a/span'
        NextPage = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NextPage)))
        NextPage.click()
        time.sleep(15)
    
        # - Click on Business Unit Tab
        xpath_BUSelect = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:NewPe1:0:businessUnitId::lovIconId"]'
        BUSelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_BUSelect)))
        BUSelectClick.click()
        time.sleep(5)
        # - Select the search button
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
        time.sleep(3)
        # - Give the Business Details
        xpath_BU = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:NewPe1:0:businessUnitId::_afrLovInternalQueryId:value00::content"]'
        Business_Unit= WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_BU)))
        Business_Unit.clear()
        BusinessUnit= readData(path,"Emp Creation",r,7)
        Business_Unit.send_keys(BusinessUnit+ Keys.ENTER)
        xpath_selectBU = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:NewPe1:0:businessUnitId_afrLovInternalTableId::db"]/table/tbody/tr/td[2]/div/table/tbody/tr/td/span'
        selectBU = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectBU)))
        selectBU.click()
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
        time.sleep(10)
        
        
        # - Click on Position Tab
        xpath_PositionSelect = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:positionId::lovIconId"]'
        PostionSelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_PositionSelect)))
        PostionSelectClick.click()
        time.sleep(5)
        # - Select the search button
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
        time.sleep(3)
        # - Give the Position Details
        xpath_Position = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:positionId::_afrLovInternalQueryId:value00::content"]'
        Positon_Details= WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Position)))
        Positon_Details.clear()
        Position= readData(path,"Emp Creation",r,8)
        Positon_Details.send_keys(Position+ Keys.ENTER)
        xpath_selectPosition = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:positionId_afrLovInternalTableId::db"]/table/tbody/tr/td[2]/div/table/tbody/tr/td[1]'
        selectPosition = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectPosition)))
        selectPosition.click()
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
        time.sleep(5)
        print("Position Created")
        
        
        #Scroll down
        body = driver.find_element(By.TAG_NAME, 'body')
        body.send_keys(Keys.PAGE_DOWN) 
        print("Page_Down")
        time.sleep(2)
           
    
   #     - Click on Job Tab
        
        excel_Job_Details= readData(path,"Emp Creation",r,9)
        #//*[@id='pt1:_FOr1:1:_FONSr2:0:MAt2:0:AP1:AT1:_ATp:Lines:0:Item::content']
        xpath_Job = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:jobId::content"]'
        Job_Details = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Job)))
        Job_Details.clear()
        Job_Details.send_keys(excel_Job_Details+ Keys.ENTER)
        time.sleep(1)
        print("Succes_Job")
        time.sleep(10)
        
        
        # - Click on Department Tab
        excel_Dept_Details= readData(path,"Emp Creation",r,10)
        #//*[@id='pt1:_FOr1:1:_FONSr2:0:MAt2:0:AP1:AT1:_ATp:Lines:0:Item::content']
        xpath_Dept= '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:departmentId::content"]'
        Dept_Details = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Dept)))
        Dept_Details.clear()
        Dept_Details.send_keys(excel_Dept_Details+ Keys.ENTER)
        print("Succes_Dept")
        time.sleep(3)
       
        
        # - Clcik on Next Page
        xpath_NextPage = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:tt1:next"]/a/span'
        NextPage = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NextPage)))
        NextPage.click()
        time.sleep(5)
    
        # - Clcik on Next Page
        xpath_NextPage = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:3:SP1:tt1:next"]/a/span'
        NextPage = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NextPage)))
        NextPage.click()
        time.sleep(5)
        person_label_xpath = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:4:AP1:Perso2:0:plam1"]/td[1]/label'
        person_number_xpath = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:4:AP1:Perso2:0:plam1"]/td[2]'
    
        try:
            # Check if label exists
            label_element = driver.find_element(By.XPATH, person_label_xpath)
            
            # If found, get the person number
            person_number_element = driver.find_element(By.XPATH, person_number_xpath)
            person_number_text = person_number_element.text.strip()
            # Extract only digits
            person_number_digits = re.sub(r'\D', '', person_number_text)
            writeData(path, "Termination", r, 1, person_number_digits)
            print("Person Number:", person_number_digits)
            #print("Person Number:", person_number)
        
        except NoSuchElementException:
            print("Person Number label not found")
        
        # - Clcik on Submit
        xpath_Submit = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:4:AP1:tt1:submit"]/a/span'
        Submit = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Submit)))
        Submit.click()
        time.sleep(5)
    
        # - Click on Yes Box
        xpath_clickYes = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:4:AP1:tt1:okWarningDialog"]/span'
        clickYes = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickYes)))
        clickYes.click()
        print("Mouli")
        time.sleep(15)
    
        # - Click on ok Box
        xpath_clickOK = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:4:AP1:tt1:okConfirmationDialog"]'
        clickOK = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickOK)))
        clickOK.click()
        
        writeData(path,"Termination",r,1,person_number)
        writeData(path,"Emp Creation",r,11,"Pass")
        writeData(path,"Emp Creation",r,12,"Employee created Successfually -person Number: " + person_number)
    
        time.sleep(5)
    driver.quit()

    wb = openpyxl.load_workbook(path)
    excel_bytes = Excel.workbook_to_bytes(wb)
    
    st.download_button(
                label="Download Excel File",
                data = excel_bytes,
                file_name=f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
@mcp.tool()
def Termination_Employee (path: str = excel_path):
    
    """Termination of an Employee"""
    excel_exists = Excel.check_excel_exists(path)
    print("Hi")
    if not excel_exists:
        st.error("Excel file not found. Please check the file path.")
        return  
    Excel.close_open_excel()
    
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
    service=ChromeService(ChromeDriverManager().install()), 
    options=options
        
        )
    driver.maximize_window()
    rows_Termination = getRowCount(path, 'Termination')
          
    #path = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    rows1 = getRowCount(path,'Config')
    rows2 = getRowCount(path,'Termination')
    Base_url = readData(path,"Config",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"Config",2,2)
    PW = readData(path,"Config",2,3)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()
    # - clicking on home button
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()
    # - clicking on my client groups
    xpath_MCG = '//a[@id="groupNode_workforce_management"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()
    # - clicking on person Managment button
        
    xpath_MCG = '//a[@id="itemNode_workforce_management_person_management_0"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()
        
    #Enter person number
        # 🔑 CRITICAL FIX 2: Added +1 to the range to include the last row of data
    created = 0
    for r in range(2, rows_Termination+1):
        
        # Read the data from the 'Termination' sheet, column 1
        excel_person_number = readData(path, 'Termination', r, 1)
        print(excel_person_number)
        #"//input[@aria-label=' Person Number']"
        xpath_PersonNumber = "//input[contains(@id, 'q1:value10::content')]"
        person_number_input = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, xpath_PersonNumber))
        )
        #
        print("Test1")
        # Clear and send keys for the current person number
        #person_number_input.clear()
        time.sleep(5)
        person_number_input.send_keys(excel_person_number+ Keys.ENTER)
        time.sleep(5) 
        print("Test2")
        # Wait for the search result before proceeding to the next row
        time.sleep(5) 
        xpath_PersonNumber_Search = "//button[text()='Search']"
        
        PersonSearch = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, xpath_PersonNumber_Search))
        )
        PersonSearch.click()
        
        #click on Action button
        action_icon_xpath = "//button[@title='Actions']"
        action_button = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, action_icon_xpath))
        )
        #//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:table1:_ATp:table2:0:cil1"]/img
        action_button.click()
        #time.sleep(5)
        
        # Click Person & Employment
        xpath_Person = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:table1:am2:dc_i1:3:dcm1"]/td[2]'
        Person_emp = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, xpath_Person))
        )
        Person_emp.click()
        
        
        # Click Work Relationship
        xpath_Workforce = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:table1:am2:dc_i1:3:dci1:12:dccmi1"]/td[2]'
        Workforce = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, xpath_Workforce))
        )
        Workforce.click()
        
        # ACTIONS
        WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@title='Actions']"))
                ).click()
        time.sleep(5)
        
        
        
        try:
            # First, click the top-level Actions menu inside the Work Relationship page
            WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@title='Actions']"))
            ).click()
            time.sleep(2)

            # Attempt to find the Terminate option
            terminate_btn_xpath = "//td[text()='Terminate']"
            terminate_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, terminate_btn_xpath))
            )
            terminate_button.click()
            time.sleep(5)
            
        except Exception as e:
            # If button not found or not clickable, mark as Fail and skip to next row
            print(f"Termination button not found for row {r}. Likely already terminated.")
            writeData(path, "Termination", r, 6, "Fail")
            writeData(path, "Termination", r, 7, "Employee already Terminated or button missing")
            
            # Navigate back to Person Management or Search to reset for the next iteration
            # Depending on your UI, you might need to click 'Done' or 'Back'
            driver.get(Base_url) # Simple way to reset state
            continue 
        
        #Enter Action details from Excel
        excel_Action = readData(path, 'Termination', r, 2)
        xpath_actionDD = "//*[contains(@id, 'Action::content')]"
        actionDD = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_actionDD)))
        actionDD.send_keys(excel_Action + Keys.ENTER)
        time.sleep(5)
        
        #Notificaton Date
        excel_Notification_Date = readData(path, 'Termination', r, 3)
        xpath_NotificationDate = "//input[@aria-label='Notification Date']"
        NotificationDate = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NotificationDate)))
        NotificationDate.clear()
        #date_string = excel_Notification_Date.strftime('%d-%b-%Y')
        date_string = excel_Notification_Date.strftime('%m/%d/%y')
        #date_string = ESDe.strftime('%m/%d/%y')
        NotificationDate.send_keys(date_string + Keys.ENTER)
        
        #Enter LastWorking day Date 
        excel_Termination_Date = readData(path, 'Termination', r, 4)
        xpath_TerminationDate = "//input[@aria-label='Last Working Day']"
        TerminationDate = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_TerminationDate)))
        TerminationDate.clear()
        #date_string_Termi = excel_Termination_Date.strftime('%d-%b-%Y')
        date_string_Termi = excel_Termination_Date.strftime('%m/%d/%y')
        TerminationDate.send_keys(date_string_Termi + Keys.ENTER)
        time.sleep(5)
        #Enter Termination Date 
        #excel_Termination_Date = readData(path, 'Termination', r, 4)
        #xpath_TerminationDate = "//input[@aria-label='Termination Date']"
        #TerminationDate = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_TerminationDate)))
        #TerminationDate.clear()
        ##date_string_Termi = excel_Termination_Date.strftime('%d-%b-%Y')
        #date_string_Termi = excel_Termination_Date.strftime('%m/%d/%y')
        #TerminationDate.send_keys(date_string_Termi + Keys.ENTER)
        #time.sleep(5)
        print("Test3")
        #Scroll down
        body = driver.find_element(By.TAG_NAME, 'body')
        body.send_keys(Keys.PAGE_DOWN) 
        print("Page_Down")
        time.sleep(2)
        # select Recommand for Rehire
        #excel_Rehire = readData(path, 'Termination', r, 5)
        #print(excel_Rehire)
        #xpath_RehireRecom_Input = "//*[contains(@id, 'RehireRecom::content')]" 
        #rehireField = WebDriverWait(driver, 20).until(
        #    EC.element_to_be_clickable((By.XPATH, xpath_RehireRecom_Input))
        #)
        #rehireField.clear()    
        #rehireField.send_keys(excel_Rehire + Keys.ENTER)
        #time.sleep(3)
        print("Succes_Rehire")
        
        # select Recommand for Rehire
        excel_Rehire = readData(path, 'Termination', r, 5)
        xpath_RehireRecom_Input = "//*[contains(@id, 'RehireRecom::content')]" 
        rehireField = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, xpath_RehireRecom_Input))
        )    
        rehireField.send_keys(excel_Rehire + Keys.ENTER)
        time.sleep(5)
        #Click on Review
        xpath_Review= '//button[text()="Review"]'
        Review = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Review)))
        Review.click()
        print("Clicked on Review")
        #time.sleep(10)
        #Submit
        #xpath_Submit1 = "//span[normalize-space(.)='Submit']"
        xpath_Submit1 = "//span[@class='xrk' and contains(., 'Submit')]"
        print("Clicked on Submit1")
        xpath_Submit1 = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt1:1:r1:0:r1:1:pt1:sp1:tt1:submit"]/a/span'
        print("Clicked on Submit1")
        Submit = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Submit1)))
        Submit.click()
        print("Clicked on Submit2")
        time.sleep(5)
                    
        #Click on Yes in Warning Box
        
        #//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt2:1:r1:0:r1:1:pt1:sp1:tt1:okWarningDialog"]
        #
        xpath_Click_Yes= "//button[@accesskey='Y']"
        Click_Yes = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Click_Yes)))
        Click_Yes.click()
        print("Clicked on Yes in Warning Box")
        time.sleep(5)
        
        #
        ##Click on Ok in Warnming Box
        #
        xpath_Click_Ok= "//button[@accesskey='K']"
        Click_Ok = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Click_Ok)))
        Click_Ok.click()
        print("Clicked on Ok in Confirmation Box")
        writeData(path,"Termination",r,6,"Pass")
        writeData(path,"Termination",r,7,"Employee Terminated Successfually")
    driver.quit()

    wb = openpyxl.load_workbook(path)
    excel_bytes = Excel.workbook_to_bytes(wb)
    
    st.download_button(
                 label="Download Excel File",
                 data = excel_bytes,
                file_name=f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
@mcp.tool()
def BusinessUnit_Creation (path: str = excel_path):
    """Business Unit Creation"""
    excel_exists = Excel.check_excel_exists(path)
    #print("Hi_BU")
    if not excel_exists:
        st.error("Excel file not found. Please check the file path.")
        return  
    Excel.close_open_excel()
    #print(path)
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()), 
        options=options
    )
    driver.maximize_window()
    #rows_Termination = getRowCount(path, 'BU')
    #print("Hi_BU2 crossed chrome")      
    #path = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    rows1 = getRowCount(path,'Config')
    rows2 = getRowCount(path,'BU')
    Base_url = readData(path,"Config",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"Config",2,2)
    PW = readData(path,"Config",2,3)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()
    
    # - clicking on home button
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()

    # - clicking on my client groups
    xpath_MCG = '//a[@id="groupNode_workforce_management"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()

    # - clicking on workforce structure
    xpath_WS = '//*[@id="itemNode_workforce_management_workforce_structures_0"]'
    WSClick = WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH, xpath_WS)))
    WSClick.click()

    # - clicking on manage business units
    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Manage Business Units")))
    MBlink = driver.find_element(By.PARTIAL_LINK_TEXT, "Manage Business Units")
    MBlink.click()

    # - process starts from clicking in add
    for r in range (2,rows2+1):
        time.sleep(2)

        xpath_add = "//img[contains(@id, 'ATp:create::icon')]"
        addClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_add)))
        addClick.click()

        time.sleep(2)
        
        xpath_name = "//input[contains(@id, 'pnlProp:inputText3::content')]"
        name = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_name)))
        name.clear()
        nameE = readData(path,"BU",r,1)
        name.send_keys(nameE+Keys.ENTER)

        xpath_setdd = "//a[contains(@id, 'defaultSetCodeId::lovIconId')]"
        setdd = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_setdd)))
        setdd.click()

        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        MBlink = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        MBlink.click()

        xpath_RDSN = "//input[@aria-label=' Reference Data Set Name']"
        RDSN_SB = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_RDSN)))
        RDSN_SB.clear()
        RDS = readData(path,"BU",r,2)
        RDSN_SB.send_keys(RDS+ Keys.ENTER)

        time.sleep(3)

        xpath_selectRDSN = "//td[@_afrrh='true']"
        selectRDSN = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectRDSN)))
        selectRDSN.click()

        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
        time.sleep(2)

        # 1. Click Save and Close
        xpath_saveclose = "//a[./span[text()='Save and Close']]"
        saveclose = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_saveclose)))
        saveclose.click()

        # 2. Check for Error Popup
        try:
            xpath_clickOK = "//button[text()='OK']"
            error_ok_btn = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH, xpath_clickOK)))
            
            error_ok_btn.click()
            time.sleep(1)
            
            xpath_cancel = "//a[@accesskey='C']"
            cancel_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath_cancel)))
            driver.execute_script("arguments[0].click();", cancel_button)
            
            time.sleep(2)
            writeData(path,"BU",r,3,"Fail")
            writeData(path,"BU",r,4,"Business Unit creation failed - Duplicate or Error")   
        except TimeoutException:
            WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, "//img[contains(@id, 'ATp:create::icon')]")))
            writeData(path,"BU",r,3,"Pass")
            writeData(path,"BU",r,4,"Business Unit created successfully")

    time.sleep(3)
    driver.quit()
    wb = openpyxl.load_workbook(path)
    excel_bytes = Excel.workbook_to_bytes(wb)
    
    st.download_button(
                 label="Download Excel File",
                 data = excel_bytes,
                file_name=f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
@mcp.tool()
def PO_Creation (path: str = excel_path):
    """Purchase Order Creation"""
    excel_exists = Excel.check_excel_exists(path)
    print("Hi_PO")  
    if not excel_exists:
        st.error("Excel file not found. Please check the file path.")
        return  
    Excel.close_open_excel()
    #print(path)
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()), 
        options=options
    )
    driver.maximize_window()
    rows1 = getRowCount(path,'Config')
    rows2 = getRowCount(path,'PO')
    Base_url = readData(path,"Config",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"Config",2,2)
    PW = readData(path,"Config",2,3)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()  
   


    # Implementation of PO Creation goes here   
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()
 
    RIGHT_NAV_LOCATOR = (By.ID, "clusters-right-nav")
 
    RightNavClick = WebDriverWait(driver, 15).until( EC.element_to_be_clickable(RIGHT_NAV_LOCATOR) )
 
    num_clicks = 6
 
    for i in range(num_clicks):
 
        RightNavClick.click()
 
    print(driver)    
    time.sleep(5)
    PROCUREMENT_LINK = (By.XPATH, '//a[@id="groupNode_procurement"]') 
    ProcurementClick = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(PROCUREMENT_LINK))
    ProcurementClick.click()
    PURCHASE_ORDERS_LINK = (By.XPATH, '//*[@id="itemNode_procurement_PurchaseOrders"]')

    PurchaseOrdersClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable(PURCHASE_ORDERS_LINK))

    PurchaseOrdersClick.click()
    rows = getRowCount(path,'PO')
    for r in range (2,rows+1):

        TASKS_ICON = (By.XPATH, '//*[@id="pt1:_FOr1:1:_FONSr2:0:_FOTsdi__PrcPoPurchaseOrdersWorkarea_itemNode__FndTasksList::icon"]')
        selecttaskbar = WebDriverWait(driver, 15).until(EC.element_to_be_clickable(TASKS_ICON))
        selecttaskbar.click()

        CREATE_ORDER_LINK = (By.XPATH, '//*[@id="pt1:_FOr1:1:_FONSr2:0:_FOTRaT:0:RAtl5"]')
        selectCreateOrder = WebDriverWait(driver, 15).until(EC.element_to_be_clickable(CREATE_ORDER_LINK))
        selectCreateOrder.click()
        time.sleep(5)
    
        xpath_POStyle = '//*[@id="pt1:_FOr1:1:_FONSr2:0:_FOTRaT:0:dynam1:0:StyleName::content"]'
        POStyle = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_POStyle)))
        POStyle.click()
        POStyle.clear()
        POStyle1 = readData(path,'PO',r,1)
        POStyle.send_keys(POStyle1 + Keys.ENTER)
        time.sleep(2)
    
        xpath_ProcBU = '//*[@id="pt1:_FOr1:1:_FONSr2:0:_FOTRaT:0:dynam1:0:ProcurementBu::content"]'
        ProcBU = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ProcBU)))
        ProcBU.click()
        ProcBU1 = readData(path,'PO',r,2)
        ProcBU.send_keys(ProcBU1 + Keys.ENTER)
        time.sleep(2)

        xpath_Supplier= '//*[@id="pt1:_FOr1:1:_FONSr2:0:_FOTRaT:0:dynam1:0:Supplier::content"]'
        Supplier = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Supplier)))
        Supplier.click()
        Supplier1 = readData(path,'PO',r,3)
        Supplier.send_keys(Supplier1 + Keys.ENTER)
        time.sleep(2)
    
        xpath_Create = '//*[@id="pt1:_FOr1:1:_FONSr2:0:_FOTRaT:0:dynam1:0:commandButton1"]'
        Create = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Create)))
        Create.click()
        time.sleep(2)


        wait = WebDriverWait(driver, 15)
        xpath_createLines = "//a[contains(@id,'lineDetailItemId::disAcr')]"
        createLines = wait.until(EC.presence_of_element_located((By.XPATH, xpath_createLines)))
        driver.execute_script("""
            arguments[0].scrollIntoView(true);
            arguments[0].click();
        """, createLines)
        print("Lines section opened")

        xpath_AddRows = "//img[@title='Add Row']"
        AddRows = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_AddRows)))
        AddRows.click()
        time.sleep(5)
        print("Tesst2")

        xpath_AddDescrip = "//textarea[contains(@id,'ItemDescription::content')]"
        AddDescrip = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, xpath_AddDescrip)))
        driver.execute_script("arguments[0].scrollIntoView(true);", AddDescrip)
        AddDescrip.click()
        AddDescrip1 = readData(path,'PO',r,4)
        AddDescrip.send_keys(AddDescrip1 + Keys.ENTER)
        time.sleep(2)

        xpath_CatSearch = "//input[contains(@id,'Category::content')]"
        CatSearch = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_CatSearch)))
        CatSearch1 = readData(path,'PO',r,5)
        CatSearch.send_keys(CatSearch1 + Keys.ENTER)
        time.sleep(2)


        xpath_AddQuantity = "//input[contains(@id,'Quantity::content')]"

    def set_adf_quantity(driver, value, retries=3):
        actions = ActionChains(driver)

        for attempt in range(retries):
            try:
                # 🔁 ALWAYS re-locate (ADF re-renders rows)
                qty = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, xpath_AddQuantity))
                )

                # 🔥 HARD SCROLL (ADF container scroll)
                actions.move_to_element(qty).perform()

                driver.execute_script("""
                    arguments[0].scrollIntoView({block:'center'});
                """, qty)

                time.sleep(0.6)

                # 🔥 ACTIVATE EDIT MODE (THIS IS THE KEY)
                driver.execute_script("""
                    arguments[0].click();
                    arguments[0].focus();
                """, qty)

                time.sleep(0.4)

                # ✅ CLEAR (ADF-safe)
                driver.execute_script("""
                    arguments[0].value = '';
                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                """, qty)

                # ✅ SET VALUE FROM EXCEL
                driver.execute_script("""
                    arguments[0].value = arguments[1];
                    arguments[0].dispatchEvent(new Event('input',  { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                """, qty, str(value))

                # 🔑 COMMIT VALUE (MANDATORY)
                qty.send_keys(Keys.TAB)

                return  # ✅ SUCCESS

            except StaleElementReferenceException:
                print(f"⚠️ Quantity stale, retrying ({attempt + 1})...")
                time.sleep(1)

        raise Exception("❌ Unable to set Quantity after retries")
    AddQuantity1 = readData(path, 'PO', r, 6)
    set_adf_quantity(driver, AddQuantity1)
    print("✅ Quantity entered successfully")




    xpath_uom = "//input[contains(@id,'Uom::content')]"

    for attempt in range(3):
        try:
            uom = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.XPATH, xpath_uom))
            )

            uom1 = readData(path, 'PO', r, 7)

            driver.execute_script("""
                arguments[0].scrollIntoView(true);
                arguments[0].removeAttribute('readonly');
                arguments[0].focus();
                arguments[0].value = arguments[1];
                arguments[0].dispatchEvent(new Event('input',  { bubbles: true }));
                arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                arguments[0].dispatchEvent(new Event('blur',   { bubbles: true }));
            """, uom   , str(uom1))

            uom.send_keys(Keys.TAB)   # 🔑 mandatory for ADF
            break

        except Exception as e:
            print("Retrying UOM due to:", e)
            time.sleep(6)
    print("open")


    xpath_Price = "//input[contains(@id,'ListPriceGoods::content')]"

    def set_adf_price(driver, value="500.00", retries=3):
        actions = ActionChains(driver)

        for attempt in range(retries):
            try:
                # 🔁 ALWAYS re-locate (ADF re-renders rows)
                price = WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((By.XPATH, xpath_Price))
                )

                # 🔥 HARD SCROLL (ADF container + mouse)
                actions.move_to_element(price).perform()

                driver.execute_script("""
                    arguments[0].scrollIntoView({block:'center', inline:'nearest'});
                """, price)

                time.sleep(0.6)

                # 🔥 Activate edit mode (VERY IMPORTANT)
                driver.execute_script("""
                    arguments[0].click();
                    arguments[0].focus();
                """, price)

                time.sleep(0.5)

                # ✅ Clear value (ADF-safe)
                driver.execute_script("""
                    arguments[0].value = '';
                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                """, price)

                # ✅ Set value
                driver.execute_script("""
                    arguments[0].value = arguments[1];
                    arguments[0].dispatchEvent(new Event('input',  { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                """, price, value)

                # 🔑 Commit value (MANDATORY for ADF)
                price.send_keys(Keys.TAB)

                return  # ✅ SUCCESS

            except StaleElementReferenceException:
                print(f"⚠️ Price stale, retrying ({attempt + 1})...")
                time.sleep(1)

        raise Exception("❌ Unable to set Price after multiple retries")
    price_value = readData(path, 'PO', r, 8)
    set_adf_price(driver, str(price_value))
    print("✅ Price entered and committed successfully")

    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.common.exceptions import StaleElementReferenceException

    xpath_ReqDate = "//input[contains(@id,'NeedByDate::content')]"  

    def set_adf_date(driver, date_value, retries=3):
        actions = ActionChains(driver)

        for attempt in range(retries):
            try:
                ReqDate = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((By.XPATH, xpath_ReqDate))
                )

                # 🔥 Scroll inside ADF container
                actions.move_to_element(ReqDate).perform()
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", ReqDate
                )

                time.sleep(0.5)

                # 🔥 Activate edit mode
                driver.execute_script("""
                    arguments[0].removeAttribute('readonly');
                    arguments[0].click();
                    arguments[0].focus();
                """, ReqDate)

                time.sleep(0.3)

                # ✅ Clear existing value
                driver.execute_script("""
                    arguments[0].value = '';
                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                """, ReqDate)

                # ✅ Set date from Excel
                driver.execute_script("""
                    arguments[0].value = arguments[1];
                    arguments[0].dispatchEvent(new Event('input',  { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('blur',   { bubbles: true }));
                """, ReqDate, date_value)

                # 🔑 Commit value (MANDATORY in ADF)
                ReqDate.send_keys(Keys.TAB)

                return  # ✅ SUCCESS

            except StaleElementReferenceException:
                print(f"⚠️ Date stale, retrying ({attempt + 1})...")
                time.sleep(1)

        raise Exception("❌ Unable to set Need By Date")
    reqdate = readData(path, 'PO', r, 9)
    set_adf_date(driver, str(reqdate))
    print("✅ Need By Date entered successfully")
    time.sleep(6)

    xpath_Actions = "//a[contains(@id,'::popEl') and @title='Actions']"
    Actions = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.XPATH, xpath_Actions))
    )
    Actions.click()

    # Wait for ADF overlay to disappear
    WebDriverWait(driver, 60).until(
        EC.invisibility_of_element_located((By.CLASS_NAME, "AFModalGlassPane"))
    )

    xpath_Validate = "//*[(@role='menuitem' or self::td or self::a)]//*[contains(normalize-space(),'Validate')]"
    Validate = WebDriverWait(driver, 60).until(
        EC.presence_of_element_located((By.XPATH, xpath_Validate))
    )

    driver.execute_script("arguments[0].click();", Validate)
    xpath_Submit = "//span[normalize-space(.)='Submit']"
    Submit = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Submit)))
    Submit.click()

    writeData(path,"PO",r,10,"Pass")
    writeData(path,"PO",r,11,"PO Created Successfully")
    time.sleep(10)
    driver.quit()
    
    wb = openpyxl.load_workbook(path)
    excel_bytes = Excel.workbook_to_bytes(wb)
    st.download_button(
                 label="Download Excel File",
                 data = excel_bytes,
                file_name=f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

def do_call_tool():
    try:
        print(tool_name)
        if tool_name == "Department_creation":
            Manage_Departments()
           
        elif tool_name == "Job_creation":
            Manage_Jobs()        
        
        elif tool_name == "Position_creation":
            Manage_Positions()   
        
        elif tool_name == "Employee_creation":
            Employee_Creation()
        
        elif tool_name == "Employee_Termination":
            Termination_Employee()
        elif tool_name == "Business_Unit_Creation":
            BusinessUnit_Creation()
        elif tool_name == "Purchase_Order_Creation":
            PO_Creation()

        else:
            #Server.py
            result = type('Result', (object,), {'content': f"No call executed. Tool: {tool_name}"})

        
    except Exception as e:
        st.session_state.chat.append({"role": "server", "text": f"Error calling tool: {e}"})

# Button handling (Streamlit re-runs)
if list_button:
    do_list_tools()
if call_button:
    do_call_tool()

if __name__ == "__main__":
    mcp.run(transport="stdio")
   #mcp.run(transport="streamable-http")
   # mcp.run(transport="http", host="0.0.0.0", port=8000)
