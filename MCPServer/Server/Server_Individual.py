from dotenv import load_dotenv
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException,NoSuchElementException
import re
from mcp.server.fastmcp import FastMCP



load_dotenv("../.env")

# Create an MCP server
mcp = FastMCP(name="Test_Automation_Individual",host="0.0.0.0",port=8050)

#mcp = FastMCP(name="Test_Automation",port=8000)
print("HCM_Automation Module Loaded_Invidual")
#To get row count#
def getRowCount(file: str, sheetName: str) -> int:
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb[sheetName]
    count = sheet.max_row
    wb.close()
    return count

#To get Column count#
def getColumnCount(file: str, sheetName: str) -> int:
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb[sheetName]
    count = sheet.max_column
    wb.close()
    return count

#To Read Data from Excel#
def readData(file: str, sheetName: str, rownum: int, columnno: int):
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb[sheetName]
    value = sheet.cell(row=rownum, column=columnno).value
    wb.close()
    return value

#To write Data to Excel#
def writeData(file: str, sheetName: str, rownum: int, columno: int, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    sheet.cell(row=rownum, column=columno).value = data
    wb.save(file)
    wb.close()

@mcp.tool()
def Manage_Jobs (Path: str = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"):
    """Job Creation"""
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
    service=ChromeService(ChromeDriverManager().install()), 
    options=options   
        )
    driver.maximize_window()
    path = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    rows1 = getRowCount(path,'Config')
    rows2 = getRowCount(path,'Job Creation')
    Base_url = readData(path,"Config",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"Config",2,2)
    PW = readData(path,"Config",2,3)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()
    
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()
    
    # - clicking on settings&actions, then clicking on setup & maintenance, clicking on taskbar
    
    xpath_ic = "//img[@title='Settings and Actions']"
    ic = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ic)))
    ic.click()
    
    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Setup and Maintenance")))
    SMlink = driver.find_element(By.PARTIAL_LINK_TEXT, "Setup and Maintenance")
    SMlink.click()
    
    xpath_Tasks = "//img[@title='Tasks']"
    Tasks = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Tasks)))
    Tasks.click()
    
    # - Clicking on Search -> searching managa job and clicking on it
    
    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
    Searchlink1 = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
    Searchlink1.click()
    
    xpath_ITF = "//input[contains(@id, ':s9:it1::content')]"
    ITF = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ITF)))
    ITF.click()
    ITF.send_keys("MANAGE JOB"+ Keys.ENTER)
    
    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Manage Job")))
    MPLink = driver.find_element(By.PARTIAL_LINK_TEXT, "Manage Job")
    MPLink.click()
    
    for r in range (2,rows2+1):
        xpath_ADD = "//img[contains(@src, 'func_add_16_ena.png')]"
        ADD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ADD)))
        ADD.click()
    
        xpath_ESD = "//input[@aria-label='Effective Start Date']"
        ESD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ESD)))
        ESD.click()
        ESD.clear()
        ESDe = readData(path,"Job Creation",r,1)
        date_string = ESDe.strftime('%d-%b-%Y')
        ESD.send_keys(date_string)
    
        xpath_jobDD = "//a[@title='Search: Job Set']"
        jobDD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobDD)))
        jobDD.click()
        element_to_click = WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        driver.execute_script("arguments[0].scrollIntoView();", element_to_click)
        SLink1 = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        SLink1.click()
 
    
        xpath_RDSN = "//input[@aria-label=' Reference Data Set Name']"
        RDSN_SB = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_RDSN)))
        RDSN_SB.clear()
        RDS = readData(path,"Job Creation",r,2)
        RDSN_SB.send_keys(RDS+ Keys.ENTER)
    
        xpath_selectRDSN = '//*[@id="pt1:r1:0:rt:1:r2:0:dynamicRegion1:1:basicDetailsPnl:setName_afrLovInternalTableId::db"]/table/tbody/tr/td[1]'
        selectRDSN = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectRDSN)))
        selectRDSN.click()
    
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
    
        xpath_name = "//input[contains(@id, ':basicDetailsPnl:name::content')]"
        name = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_name)))
        name.clear()
        nameE = readData(path,"Job Creation",r,3)
        name.send_keys(nameE+Keys.ENTER)
    
        xpath_jobcode = "//input[contains(@id, ':basicDetailsPnl:jobCode::content')]"
        jobcode = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobcode)))
        jobcode.clear()
        JCE = readData(path,"Job Creation",r,4)
        jobcode.send_keys(JCE+Keys.ENTER)
        time.sleep(3)
    
        xpath_error_icon = "//td[contains(text(), '(PER-1530038)')]"
        xpath_clickCE = "//a[normalize-space(.)='Cancel']"
        try :
            error_icon = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath_error_icon)))
            cancel_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath_clickCE)))
            driver.execute_script("arguments[0].click();", cancel_button)
            xpath_clickYE_confirmation = "//button[@accesskey='Y']" 
            try:
                #- If it appears, clicks it.
                WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, xpath_clickYE_confirmation))).click()
                time.sleep(1) 
            except TimeoutException:
                pass 
    
            writeData(path,"Job Creation",r,6,"code EXISTS")
            time.sleep(3)
            continue
        except TimeoutException:
            pass        
    
        xpath_next1 = "//a[@accesskey='x']"
        next1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next1)))
        next1.click()
        time.sleep(3)
    
        xpath_AStatus = "//input[@title='Active']"
        AStatus = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_AStatus)))
        AStatus.click()
        AStatusE = readData(path,"Job Creation",r,5)
        AStatus.send_keys(AStatusE+ Keys.ENTER)
    
        xpath_next2 = "//a[@accesskey='x']"
        next2 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next2)))
        next2.click()
        time.sleep(3)
    
        xpath_next3 = "//a[@accesskey='x']"
        next3 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next3)))
        next3.click()
        time.sleep(3)
    
        xpath_next4 = "//a[@accesskey='x']"
        next4 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next4)))
        next4.click()
        time.sleep(3)
    
        xpath_submit = "//a[normalize-space(.)='Submit']"
        submit = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_submit)))
        submit.click()
        time.sleep(2)
    
        xpath_yes1 = "//button[@accesskey='Y']"
        yes1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_yes1)))
        yes1.click()
        time.sleep(3)
    
        driver.switch_to.active_element.send_keys(Keys.TAB)
        driver.switch_to.active_element.send_keys(Keys.ENTER)
    
        writeData(path,"Job Creation",r,6,"JOB CREATED")
        
    
        time.sleep(3)
    

    driver.quit()

@mcp.tool()

def Manage_Departments (path: str = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"):
    """Department Creation"""
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
    service=ChromeService(ChromeDriverManager().install()), 
    options=options
        
        )
    driver.maximize_window()
          
    path = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    rows1 = getRowCount(path,'Config')
    rows2 = getRowCount(path,'Dept Creation')
    Base_url = readData(path,"Config",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"Config",2,2)
    PW = readData(path,"Config",2,3)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()
    
    # - clicking on home button
    
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()
    
    # - clicking on my client groups
    
    xpath_MCG = '//a[@id="groupNode_workforce_management"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()
    
    # - clicking on workforce structure
    
    xpath_WS = '//*[@id="itemNode_workforce_management_workforce_structures_0"]'
    WSClick = WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH, xpath_WS)))
    WSClick.click()
    
    # - searching for manage department and clicking on it
    
    #xpath_SB = "//input[@placeholder='Search for tasks']"
    #searchBox = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_SB)))
    #searchBox.send_keys("manage department" + Keys.ENTER)
    
    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Manage Departments")))
    link = driver.find_element(By.PARTIAL_LINK_TEXT, "Manage Departments")
    link.click()
    
    for r in range (2,rows2+1):
        xpath_create = "//a[./span[text()='Create']]"
        createClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_create)))
        createClick.click()
    
        xpath_calender = "//input[@aria-label='Effective Start Date']"
        searchBox = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_calender)))
        searchBox.clear()
        date = readData(path,"Dept Creation",r,1)
        date_string = date.strftime('%d-%b-%Y')
        searchBox.send_keys(date_string)
    
        xpath_deptSelect = "//a[@title='Search: Department Set']"
        deptSelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_deptSelect)))
        deptSelectClick.click()
    
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
    
        xpath_RDSN = "//input[@aria-label=' Reference Data Set Name']"
        RDSN_SB = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_RDSN)))
        RDSN_SB.clear()
        RDS = readData(path,"Dept Creation",r,2)
        RDSN_SB.send_keys(RDS+ Keys.ENTER)
    
        xpath_selectRDSN = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:AP1:setName2Id_afrLovInternalTableId::db"]/table/tbody/tr/td[1]'
        selectRDSN = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectRDSN)))
        selectRDSN.click()
    
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
        time.sleep(2)
    
        NM = readData(path,"Dept Creation",r,3)
        driver.switch_to.active_element.send_keys(Keys.TAB + NM)
    
        xpath_statusDD = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:2:AP1:selectOneChoice2::content"]'
        statusDD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_statusDD)))
        statusDD.click()
    
        xpath_Astatus = "//li[text()='Active']"
   
        # 1. Click 'Active' status (This action triggers the Name Exists validation)
        Astatus = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_Astatus)))
        Astatus.click()
        time.sleep(2)
 
        # - START DIALOG CHECK BLOCK
        # - We use a try/except to check for the MANDATORY presence of the error dialog button (OK)
        # - If the OK button is found, we assume the error occurred and execute the 'except' block.
   
        xpath_error_ok_button = "//button[text()='OK']"
   
        try:    
       
            error_ok_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_error_ok_button)))
            error_ok_button.click()
            time.sleep(1)
 
            xpath_clickCE = "//a[normalize-space(.)='Cancel']"
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath_clickCE))).click()
            time.sleep(1)
 
            xpath_clickYE_confirmation = "//button[@accesskey='Y']"
            try:
                #- If it appears, clicks it.
                WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, xpath_clickYE_confirmation))).click()
                time.sleep(1)
            except TimeoutException:
                pass
 
            writeData(path,"Dept Creation",r,4,"NAME EXISTS")
            time.sleep(3)
            continue
           
        except TimeoutException:
        #- IF DIALOG DID NOT APPEAR
               
            xpath_clickNext = "//a[normalize-space(.)='Next']"
            clickNext = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickNext)))
            clickNext.click()
 
            xpath_clickNext2 = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:3:AP2:tt1:next"]/a/span'
            clickNext2 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickNext2)))
            clickNext2.click()
 
            xpath_clickSubmit = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:4:ap1:tt1:submit"]/a/span'
            clickSubmit = WebDriverWait(driver,15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickSubmit)))
            clickSubmit.click()
 
            xpath_clickYes = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:4:ap1:tt1:okWarningDialog"]'
            clickYes = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickYes)))
            clickYes.click()
 
            xpath_clickOK = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:4:ap1:tt1:okConfirmationDialog"]'
            clickOK = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickOK)))
            clickOK.click()
 
            writeData(path,"Dept Creation",r,4,"DEPT CREATED")
            time.sleep(2)
 
    driver.quit()

@mcp.tool()
def Manage_Positions (path: str = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"):
    """Position Creation"""
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), 
    options=options  )
    driver.maximize_window()
    rows_Termination = getRowCount(path, 'Pos Creation')
          
    path = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    rows1 = getRowCount(path,'Config')
    rows2 = getRowCount(path,'Pos Creation')
    Base_url = readData(path,"Config",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"Config",2,2)
    PW = readData(path,"Config",2,3)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()   

    # - clicking on home button

    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()

    # - clicking on settings&actions, then clicking on setup & maintenance, clicking on taskbar

    xpath_ic = "//img[@title='Settings and Actions']"
    ic = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ic)))
    ic.click()

    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Setup and Maintenance")))
    SMlink = driver.find_element(By.PARTIAL_LINK_TEXT, "Setup and Maintenance")
    SMlink.click()

    xpath_Tasks = "//img[@title='Tasks']"
    Tasks = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Tasks)))
    Tasks.click()

    # - Clicking on Search -> searching managa positions and clicking on it

    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
    Searchlink1 = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
    Searchlink1.click()

    xpath_ITF = "//input[contains(@id, ':s9:it1::content')]"
    ITF = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ITF)))
    ITF.click()
    ITF.send_keys("MANAGE POSITIONS"+ Keys.ENTER)

    WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Manage Positions")))
    MPLink = driver.find_element(By.PARTIAL_LINK_TEXT, "Manage Positions")
    MPLink.click()

    #xpath_ADD = "//input[contains(@id, ':s9:it1::content')]"
    #ITF = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ITF)))
    #ITF.click()

    # - clicking on ADD and entering details in 1st BASIC DETAILS page
    for r in range (2,rows2+1):

        xpath_ADD = "//img[contains(@src, 'func_add_16_ena.png')]"
        ADD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ADD)))
        ADD.click()

        xpath_ESD = "//input[@aria-label='Effective Start Date']"
        ESD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_ESD)))
        ESD.click()
        ESD.clear()
        ESDe = readData(path,"Pos Creation",r,1)
        date_string = ESDe.strftime('%d-%b-%Y')
        ESD.send_keys(date_string)

        xpath_BU = "//input[contains(@id, ':AP4:BUNameId::content')]"
        BU = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_BU)))
        BU.click()
        BUe = readData(path,"Pos Creation",r,2)
        BU.send_keys(BUe+Keys.ENTER)
        time.sleep(2)

        xpath_name1 = "//input[contains(@id, ':AP4:inputText2::content')]"
        name1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_name1)))
        name1.click()
        nameE = readData(path,"Pos Creation",r,3)
        name1.send_keys(nameE+Keys.ENTER)
        time.sleep(2)

        xpath_next1 = "//a[@accesskey='x']"
        next1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next1)))
        next1.click()

        # - entering details POSITION DETAILS page

        xpath_AStatus = "//input[@title='Active']"
        AStatus = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_AStatus)))
        AStatus.click()
        AStatusE = readData(path,"Pos Creation",r,4)
        AStatus.send_keys(AStatusE+Keys.ENTER)

        xpath_deptDD = "//a[@title='Search: Department']"
        deptDD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_deptDD)))
        deptDD.click()
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        SLink1 = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        SLink1.click()

        xpath_deptN = "//input[@aria-label=' Name']"
        deptN = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_deptN)))
        deptN.clear()
        deptNE = readData(path,"Pos Creation",r,5)
        deptN.send_keys(deptNE+Keys.ENTER)
        time.sleep(3)
        xpath_selectDept = "//td[@style='width:6px;' and contains(@class, 'xwn')]"
        selectDept = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectDept)))
        selectDept.click()
        xpath_deptOK = "//button[text()='OK']"
        deptOK = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_deptOK)))
        deptOK.click()
        time.sleep(3)

        xpath_jobDD = "//a[@title='Search: Job']"
        jobDD = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobDD)))
        jobDD.click()
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        SLink1 = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        SLink1.click()
        xpath_jobName = "//input[@aria-label=' Name']"
        jobName = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobName)))
        jobName.clear()
        jobNameE = readData(path,"Pos Creation",r,6)
        jobName.send_keys(jobNameE+Keys.ENTER)
        time.sleep(2)
        xpath_jobSelect = "//td[@style='width:6px;' and contains(@class, 'xwn')]"
        jobSelect = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobSelect)))
        jobSelect.click()
        xpath_jobOK = "//button[text()='OK']"
        jobOK = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_jobOK)))
        jobOK.click()
        time.sleep(2)


        xpath_HS = "//input[contains(@id, ':selectOneChoice4::content') and @class='x2h']"
        HS = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_HS)))
        HS.click()
        HSe = readData(path,"Pos Creation",r,7)
        HS.send_keys(HSe+Keys.ENTER)
        time.sleep(2)

        xpath_type = "//a[contains(@id, ':selectOneChoice5::drop') and @class='x1kt']"
        type = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_type)))
        type.click()
        typeE = readData(path,"Pos Creation",r,8)
        type.send_keys(typeE+Keys.ENTER)
        time.sleep(2)

        xpath_HC = "//input[contains(@id, ':AP5:inputText7::content')]"
        HC = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_HC)))
        HC.clear()
        HCe = readData(path,"Pos Creation",r,9)
        HC.send_keys(HCe)
        time.sleep(2)

        xpath_next2 = "//a[normalize-space(.)='Next']"
        next2 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next2)))
        next2.click()
        time.sleep(2)
        xpath_next3 = "//a[normalize-space(.)='Next']"
        next3 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next3)))
        next3.click()
        time.sleep(2)

        xpath_next4 = "//a[normalize-space(.)='Next']"
        next4 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_next4)))
        next4.click()
        time.sleep(2)

        xpath_submit = "//a[normalize-space(.)='Submit']"
        submit = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_submit)))
        submit.click()
        time.sleep(2)

        xpath_yes1 = "//button[@accesskey='Y']"
        yes1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_yes1)))
        yes1.click()
        time.sleep(5)

        driver.switch_to.active_element.send_keys(Keys.TAB)
        driver.switch_to.active_element.send_keys(Keys.ENTER)

        writeData(path,"Pos Creation",r,10,"PASSED")

    time.sleep(3)
    driver.quit() 

@mcp.tool()
def Employee_Creation (path: str = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"):
    """Creation of an Employee"""
    
# ---------------------------------------------------------------------
#  LAUNCH BROWSER
# ---------------------------------------------------------------------
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()), 
        options=options
        )
    driver.maximize_window()
    
    # ---------------------------------------------------------------------
    #  EXCEL PATH & SHEET ROWS
    # ---------------------------------------------------------------------
   # path = "C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    rows2 = getRowCount(path, 'Emp Creation')
    rows_URL = getRowCount(path, 'Config')
    # ---------------------------------------------------------------------
    #  LOGIN
    # ---------------------------------------------------------------------
    Base_url = readData(path, "Config", 2, 1)
    driver.get(Base_url)
    
    UN = readData(path, "Config", 2, 2)
    PW = readData(path, "Config", 2, 3)
    
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()
    
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()
    
    # - clicking on my client groups
    
    xpath_MCG = '//a[@id="groupNode_workforce_management"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()
    time.sleep(5)
    
    # - Click on New Person 
    xpath_MCG = '//a[@id="itemNode_workforce_management_new_person_0"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()    
    time.sleep(5)
    
    for r in range (2,rows2+1):
        # - Clcik on Hire an Employee
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Hire an Employee")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Hire an Employee")
        link.click()
        
        # Hire Date 
        Hire_Date = readData(path,"Emp Creation",r,1)
        xpath_HireDate = "//input[@aria-label='Hire Date']"
        HireDate = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_HireDate)))
        HireDate.clear()
        Emp_Hire_Date = Hire_Date.strftime('%d-%b-%Y')
        HireDate.send_keys(Emp_Hire_Date)
        time.sleep(5)
        # Hire Action dropdown
    
        try:
    
            excel_HireAction = readData(path, 'Emp Creation', r, 2)
            # 1. Click the dropdown disclosure icon to open the list
            xpath_HireAction_disclose = '//a[contains(@id, "selectOneChoice1::drop")]'
    
            HireAction_disclose = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_HireAction_disclose))        )
    
            HireAction_disclose.click()
            print("Clicked Hire Action disclosure icon.")
            xpath_List_Item = f"//li[text()='{excel_HireAction}']" 
            HireAction_Option = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_List_Item)))
            HireAction_Option.click()
            print(f"Selected Hire Action: {excel_HireAction}") 
            time.sleep(2) # Allow time for the selection to register and the field to close
    
        except TimeoutException:
            print(f"Error: Hire Action '{excel_HireAction}' not found or clickable for row {r}.")
            #writeData(path, "Employee", r, 5, "Fail - Hire Action Select Timeout")
            continue
    
        except Exception as e:
            # Catch all other exceptions
            print(f"Error setting Hire Action for row {r}: {e}")
            #writeData(path, "Employee", r, 5, "Fail - Hire Action Error")
            continue
    
        # - Click On legal employer tab
        xpath_LESelect = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:selectOneChoice3::lovIconId"]'
        LESelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_LESelect)))
        LESelectClick.click()
        # - Select the search Button
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
        # - Give the legal employer Details
        xpath_LE = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:selectOneChoice3::_afrLovInternalQueryId:value00::content"]'
        Legal_Employer= WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_LE)))
        Legal_Employer.clear()
        LegalEntity= readData(path,"Emp Creation",r,3)
        Legal_Employer.send_keys(LegalEntity+ Keys.ENTER)
        xpath_selectLE = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:selectOneChoice3_afrLovInternalTableId::db"]/table/tbody/tr/td[2]/div/table/tbody/tr/td[1]'
        selectLE = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectLE)))
        selectLE.click()
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
    
        time.sleep(5)
    
        # - Enter Last Name
        Last_Name= readData(path,"Emp Creation",r,4)
        xpath_LastName = "//input[contains(@name, 'i1:4:it20')]"
        LastName = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,xpath_LastName)))
        LastName.clear()
        LastName.send_keys(Last_Name+ Keys.TAB)
    
        # - Enter Last Name
        First_Name= readData(path,"Emp Creation",r,5)
        xpath_FristName = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:NewPe1:0:pt_r1:0:r1:0:i1:5:it60::content"]'
        FirstName = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH,xpath_FristName)))
        FirstName.clear()
        FirstName.send_keys(First_Name+ Keys.TAB)
    
        # - Choose Gender
        xpath_Gender= '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:NewPe1:0:pt_r1:0:soc3::drop"]'
        Gender = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Gender)))
        Gender.click()
        time.sleep(5)
    
        # - select respective Gender
        xpath_Gender1 = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:NewPe1:0:pt_r1:0:soc3::pop"]/li[4]'
        Gender1 = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Gender1)))
        Gender1.click()
        time.sleep(3)
    
        # - Click on Next Page
        xpath_NextPage = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:0:SP1:tt1:next"]/a/span'
        NextPage = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NextPage)))
        NextPage.click()
        time.sleep(3)

        person_label_xpath = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:1:SP1:pt_r1:0:panelLabelAndMessage4"]/td[1]/label'
        person_number_xpath = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:1:SP1:pt_r1:0:panelLabelAndMessage4"]/td[2]'
        
        try:
            # Check if label exists
            label_element = driver.find_element(By.XPATH, person_label_xpath)
            
            # If found, get the person number
            person_number_element = driver.find_element(By.XPATH, person_number_xpath)
            person_number = person_number_element.text.strip()
        
            print("Person Number:", person_number)
        
        except NoSuchElementException:
            print("Person Number label not found")

            # - Click on Next Page
        xpath_NextPage = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:1:SP1:tt1:next"]/a/span'
        NextPage = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NextPage)))
        NextPage.click()
        time.sleep(15)
    
        # - Click on Business Unit Tab
        xpath_BUSelect = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:NewPe1:0:businessUnitId::lovIconId"]'
        BUSelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_BUSelect)))
        BUSelectClick.click()
        time.sleep(5)
        # - Select the search button
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
        time.sleep(3)
        # - Give the Business Details
        xpath_BU = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:NewPe1:0:businessUnitId::_afrLovInternalQueryId:value00::content"]'
        Business_Unit= WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_BU)))
        Business_Unit.clear()
        BusinessUnit= readData(path,"Emp Creation",r,6)
        Business_Unit.send_keys(BusinessUnit+ Keys.ENTER)
        xpath_selectBU = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:NewPe1:0:businessUnitId_afrLovInternalTableId::db"]/table/tbody/tr/td[2]/div/table/tbody/tr/td/span'
        selectBU = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectBU)))
        selectBU.click()
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
        time.sleep(10)
        
        
        # - Click on Position Tab
        xpath_PositionSelect = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:positionId::lovIconId"]'
        PostionSelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_PositionSelect)))
        PostionSelectClick.click()
        time.sleep(5)
        # - Select the search button
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
        time.sleep(3)
        # - Give the Position Details
        xpath_Position = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:positionId::_afrLovInternalQueryId:value00::content"]'
        Positon_Details= WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Position)))
        Positon_Details.clear()
        Position= readData(path,"Emp Creation",r,7)
        Positon_Details.send_keys(Position+ Keys.ENTER)
        xpath_selectPosition = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:positionId_afrLovInternalTableId::db"]/table/tbody/tr/td[2]/div/table/tbody/tr/td[1]'
        selectPosition = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectPosition)))
        selectPosition.click()
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
        time.sleep(5)
        print("Position Created")
        
        
        #Scroll down
        body = driver.find_element(By.TAG_NAME, 'body')
        body.send_keys(Keys.PAGE_DOWN) 
        print("Page_Down")
        time.sleep(2)
           
    
   #     - Click on Job Tab
        xpath_JobSelect='/html/body/div[2]/form/div[1]/div[2]/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div/div/div/div/div/div[1]/div/div/div/div[1]/div[1]/div/div/div[1]/div/div[1]/table/tbody/tr/td[1]/div/div/div/div/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div/div/div/div/div/span/div[2]/div[11]/div/table/tbody/tr/td/div/div/div/div[3]/div/div[1]/div[3]/div/div[2]/div/div[2]/div[1]/div/div[1]/table/tbody/tr/td[1]/table/tbody/tr[4]/td[2]/table/tbody/tr/td[1]/span/span/span/span/a[@title="Job"]'
        #driver.execute_script("arguments[0].scrollIntoView();", xpath_JobSelect)
        # xpath_JobSelect = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:jobId::lovIconId"]'
        JobSelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_JobSelect)))
        JobSelectClick.click()
        time.sleep(5)
        print("JOB CLICK")
        # - Select the search button
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
        time.sleep(3)
        print("JOB search CLICK")
        # - Give the Job Details
        #Job.clear()
        xpath_Job = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:jobId::_afrLovInternalQueryId:value00::content"]'
        Job= WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Job)))
        Job.clear()
        Job_Details= readData(path,"Emp Creation",r,8)
        Job.send_keys(Job_Details+ Keys.ENTER)
        print("Enter the JOB Details")
        xpath_selectJob = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:jobId_afrLovInternalTableId::db"]/table/tbody/tr/td[2]/div/table/tbody/tr/td[1]'
        selectJob = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectJob)))
        selectJob.click()
        print("choose JOB")
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
        print ("Click OK")
        time.sleep(10)
        
        
        # - Click on Department Tab
        xpath_DeptSelect = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:departmentId::lovIconId"]'
        DeptSelectClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_DeptSelect)))
        DeptSelectClick.click()
        time.sleep(5)
        # - Select the search button
        WebDriverWait(driver,15).until(EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, "Search")))
        link = driver.find_element(By.PARTIAL_LINK_TEXT, "Search")
        link.click()
        time.sleep(3)
        
        # - Give the Department Details
        xpath_Department = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:departmentId::_afrLovInternalQueryId:value00::content"]'
        Department_Details= WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Department)))
        #driver.execute_script("arguments[0].scrollIntoView();", Department_Details)
        Department_Details.clear()
        Department= readData(path,"Emp Creation",r,9)
        Department_Details.send_keys(Department+ Keys.ENTER)
        
        xpath_selectDepartment = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:NewPe3:0:JobDe1:0:departmentId_afrLovInternalTableId::db"]/table/tbody/tr/td[2]/div/table/tbody/tr/td'
        selectDepartment = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_selectDepartment)))
        selectDepartment.click()
        xpath_okClick = "//button[text()='OK']"
        okClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_okClick)))
        okClick.click()
        time.sleep(10)
        print("Department created")
        
        # - Clcik on Next Page
        xpath_NextPage = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:2:sP2:tt1:next"]/a/span'
        NextPage = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NextPage)))
        NextPage.click()
        time.sleep(5)
    
        # - Clcik on Next Page
        xpath_NextPage = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:3:SP1:tt1:next"]/a/span'
        NextPage = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NextPage)))
        NextPage.click()
        time.sleep(5)
        person_label_xpath = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:4:AP1:Perso2:0:plam1"]/td[1]/label'
        person_number_xpath = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:4:AP1:Perso2:0:plam1"]/td[2]'
    
        try:
            # Check if label exists
            label_element = driver.find_element(By.XPATH, person_label_xpath)
            
            # If found, get the person number
            person_number_element = driver.find_element(By.XPATH, person_number_xpath)
            person_number_text = person_number_element.text.strip()
            # Extract only digits
            person_number_digits = re.sub(r'\D', '', person_number_text)
            writeData(path, "Termination", r, 1, person_number_digits)
            print("Person Number:", person_number_digits)
            #print("Person Number:", person_number)
        
        except NoSuchElementException:
            print("Person Number label not found")
        
        # - Clcik on Submit
        xpath_Submit = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:4:AP1:tt1:submit"]/a/span'
        Submit = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Submit)))
        Submit.click()
        time.sleep(5)
    
        # - Click on Yes Box
        xpath_clickYes = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:4:AP1:tt1:okWarningDialog"]/span'
        clickYes = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickYes)))
        clickYes.click()
        print("Mouli")
        time.sleep(15)
    
        # - Click on ok Box
        xpath_clickOK = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAnt2:1:pt1:pt_r1:4:AP1:tt1:okConfirmationDialog"]'
        clickOK = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_clickOK)))
        clickOK.click()
        
        writeData(path,"Termination",r,1,person_number)
        writeData(path,"Emp Creation",r,10,"Pass")
        writeData(path,"Emp Creation",r,11,"Employee created Successfually -person Number: " + person_number)
    
        time.sleep(10)
    driver.quit()
@mcp.tool()


def Termination_Employee (path: str = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"):
    """Termination of an Employee"""
    options = webdriver.ChromeOptions()
    options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
    driver = webdriver.Chrome(
    service=ChromeService(ChromeDriverManager().install()), 
    options=options
        
        )
    driver.maximize_window()
    rows_Termination = getRowCount(path, 'Termination')
          
    path = r"C:\Users\velur\Desktop\Selenium-Python\Emp2Term.xlsx"
    rows1 = getRowCount(path,'Config')
    rows2 = getRowCount(path,'Termination')
    Base_url = readData(path,"Config",2,1)
    driver.get(Base_url)
    
    UN = readData(path,"Config",2,2)
    PW = readData(path,"Config",2,3)
    driver.find_element(By.ID, "userid").send_keys(UN)
    driver.find_element(By.ID, "password").send_keys(PW)
    driver.find_element(By.ID, "btnActive").click()
    # - clicking on home button
    xpath_home= '//a[@id="pt1:_UIShome"]/*[name()="svg"]/*[name()="g"][4]/*[name()="path"]'
    HomeClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_home)))
    HomeClick.click()
    # - clicking on my client groups
    xpath_MCG = '//a[@id="groupNode_workforce_management"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()
    # - clicking on person Managment button
        
    xpath_MCG = '//a[@id="itemNode_workforce_management_person_management_0"]'
    MCGClick = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_MCG)))
    MCGClick.click()
        
    #Enter person number
        # 🔑 CRITICAL FIX 2: Added +1 to the range to include the last row of data
    created = 0
    for r in range(2, rows_Termination+1):
        
        # Read the data from the 'Termination' sheet, column 1
        excel_person_number = readData(path, 'Termination', r, 1)
        
        xpath_PersonNumber = "//*[@id='_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:q1:value10::content']"
        
        person_number_input = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, xpath_PersonNumber))
        )
        
        # Clear and send keys for the current person number
        person_number_input.clear()
        person_number_input.send_keys(excel_person_number + Keys.ENTER)
        
        # Wait for the search result before proceeding to the next row
        time.sleep(5) 
        xpath_PersonNumber_Search = "//button[text()='Search']"
        
        PersonSearch = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, xpath_PersonNumber_Search))
        )
        PersonSearch.click()
        
        #click on Action button
        action_icon_xpath = "//button[@title='Actions']"
        action_button = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, action_icon_xpath))
        )
        #//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:table1:_ATp:table2:0:cil1"]/img
        action_button.click()
        #time.sleep(5)
        
        # Click Person & Employment
        xpath_Person = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:table1:am2:dc_i1:3:dcm1"]/td[2]'
        Person_emp = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, xpath_Person))
        )
        Person_emp.click()
        
        
        # Click Work Relationship
        xpath_Workforce = '//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt1:0:pt1:Perso1:0:SP3:table1:am2:dc_i1:3:dci1:12:dccmi1"]/td[2]'
        Workforce = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, xpath_Workforce))
        )
        Workforce.click()
        
        # ACTIONS
        WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@title='Actions']"))
                ).click()
        
        
        time.sleep(5)
        
        # TERMINATION BUTTON
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//td[text()='Terminate']"))).click()
        time.sleep(5)
        
        # time.sleep(5) # This was at the end of the original script
        
        
        #Enter Action details from Excel
        excel_Action = readData(path, 'Termination', r, 2)
        xpath_actionDD = "//*[contains(@id, 'Action::content')]"
        actionDD = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, xpath_actionDD)))
        actionDD.send_keys(excel_Action + Keys.ENTER)
        time.sleep(5)
        
        #Notificaton Date
        excel_Notification_Date = readData(path, 'Termination', r, 3)
        xpath_NotificationDate = "//input[@aria-label='Notification Date']"
        NotificationDate = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_NotificationDate)))
        NotificationDate.clear()
        date_string = excel_Notification_Date.strftime('%d-%b-%Y')
        NotificationDate.send_keys(date_string + Keys.ENTER)
        
        #Enter Termination Date 
        excel_Termination_Date = readData(path, 'Termination', r, 4)
        xpath_TerminationDate = "//input[@aria-label='Termination Date']"
        TerminationDate = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_TerminationDate)))
        TerminationDate.clear()
        date_string_Termi = excel_Termination_Date.strftime('%d-%b-%Y')
        TerminationDate.send_keys(date_string_Termi + Keys.ENTER)
        time.sleep(5)
        
        # select Recommand for Rehire
        excel_Rehire = readData(path, 'Termination', r, 5)
        xpath_RehireRecom_Input = "//*[contains(@id, 'RehireRecom::content')]" 
        rehireField = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, xpath_RehireRecom_Input))
        )    
        rehireField.send_keys(excel_Rehire + Keys.ENTER)
        time.sleep(5)
        #Click on Review
        xpath_Review= '//button[text()="Review"]'
        Review = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Review)))
        Review.click()
        
        time.sleep(10)
        #Submit
        xpath_Submit1 = "//span[normalize-space(.)='Submit']"
        Submit = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Submit1)))
        Submit.click()
        time.sleep(5)
                    
        #Click on Yes in Warning Box
        
        #//*[@id="_FOpt1:_FOr1:0:_FONSr2:0:MAt2:1:r1:0:r1:1:pt1:sp1:tt1:okWarningDialog"]
        #
        xpath_Click_Yes= "//button[@accesskey='Y']"
        Click_Yes = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Click_Yes)))
        Click_Yes.click()
        time.sleep(5)
        
        #
        ##Click on Ok in Warnming Box
        #
        xpath_Click_Ok= "//button[@accesskey='K']"
        Click_Ok = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, xpath_Click_Ok)))
        Click_Ok.click()
        writeData(path,"Termination",r,6,"Pass")
        writeData(path,"Termination",r,7,"Employee Terminated Successfually")
    driver.quit()


